"""
Document Processing Service with Advanced Chunking and Permission Assignment
Handles unstructured document processing with enterprise-grade metadata extraction
"""

from fastapi import FastAPI, HTTPException, UploadFile, File, Form, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
from typing import List, Dict, Optional, Any, Union
import os
import uuid
import io
import hashlib
from datetime import datetime
import json
import logging
from contextlib import asynccontextmanager
import asyncio

# Document processing imports
import PyPDF2
import docx2txt
import openpyxl
from bs4 import BeautifulSoup
import chardet
import magic

# Text processing
import re
import nltk
from nltk.tokenize import sent_tokenize, word_tokenize
from nltk.corpus import stopwords
import tiktoken

# Download required NLTK data
try:
    nltk.data.find('tokenizers/punkt')
    nltk.data.find('corpora/stopwords')
except LookupError:
    nltk.download('punkt', quiet=True)
    nltk.download('stopwords', quiet=True)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Models
class ChunkingConfig(BaseModel):
    """Configuration for document chunking strategy"""
    chunk_size: int = Field(default=1000, ge=100, le=5000, description="Maximum chunk size in characters")
    chunk_overlap: int = Field(default=200, ge=0, le=500, description="Overlap between chunks")
    chunking_strategy: str = Field(default="semantic", description="Chunking strategy: semantic, fixed, sentence")
    preserve_paragraphs: bool = Field(default=True, description="Try to keep paragraphs intact")
    min_chunk_size: int = Field(default=100, ge=50, le=1000, description="Minimum chunk size")

class PermissionTemplate(BaseModel):
    """Permission template for document access control"""
    template_name: str
    access_control_list: List[str] = Field(default=[])
    roles: List[str] = Field(default=[])
    classification: str = Field(default="internal")
    department: Optional[str] = None
    project_id: Optional[str] = None
    attributes: Dict[str, Any] = Field(default={})

class DocumentProcessingRequest(BaseModel):
    """Document processing request"""
    filename: str
    content_type: str
    permission_template: PermissionTemplate
    chunking_config: Optional[ChunkingConfig] = Field(default_factory=ChunkingConfig)
    extract_metadata: bool = Field(default=True, description="Extract document metadata")
    user_id: str
    processing_options: Dict[str, Any] = Field(default={})

class ProcessedChunk(BaseModel):
    """Processed document chunk with metadata"""
    chunk_id: str
    text: str
    chunk_index: int
    start_char: int
    end_char: int
    page_number: Optional[int] = None
    section_title: Optional[str] = None
    word_count: int
    token_count: int

class DocumentProcessingResponse(BaseModel):
    """Document processing response"""
    doc_id: str
    filename: str
    processing_status: str
    chunks: List[ProcessedChunk]
    metadata: Dict[str, Any]
    permission_metadata: Dict[str, Any]
    processing_stats: Dict[str, Any]
    created_at: str

# Document Processor
class DocumentProcessor:
    """Advanced document processing with chunking and metadata extraction"""
    
    def __init__(self):
        self.supported_formats = {
            'application/pdf': self._process_pdf,
            'application/msword': self._process_doc,
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document': self._process_docx,
            'application/vnd.ms-excel': self._process_excel,
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': self._process_excel,
            'text/plain': self._process_text,
            'text/html': self._process_html,
            'text/csv': self._process_csv
        }
        
        # Initialize tokenizer for token counting
        try:
            self.tokenizer = tiktoken.get_encoding("cl100k_base")
        except Exception:
            self.tokenizer = None
    
    def _detect_content_type(self, file_content: bytes, filename: str) -> str:
        """Detect content type from file content"""
        try:
            mime_type = magic.from_buffer(file_content, mime=True)
            return mime_type
        except Exception:
            # Fallback to extension-based detection
            ext = os.path.splitext(filename)[1].lower()
            ext_mapping = {
                '.pdf': 'application/pdf',
                '.doc': 'application/msword',
                '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                '.xls': 'application/vnd.ms-excel',
                '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                '.txt': 'text/plain',
                '.html': 'text/html',
                '.csv': 'text/csv'
            }
            return ext_mapping.get(ext, 'application/octet-stream')
    
    def _process_pdf(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """Process PDF document"""
        try:
            reader = PyPDF2.PdfReader(io.BytesIO(file_content))
            
            text_content = ""
            page_texts = []
            
            for page_num, page in enumerate(reader.pages):
                page_text = page.extract_text()
                page_texts.append({
                    'page_number': page_num + 1,
                    'text': page_text
                })
                text_content += f"\n--- Page {page_num + 1} ---\n{page_text}\n"
            
            metadata = {
                'total_pages': len(reader.pages),
                'page_texts': page_texts,
                'document_info': reader.metadata if reader.metadata else {},
                'author': reader.metadata.get('/Author', '') if reader.metadata else '',
                'title': reader.metadata.get('/Title', filename) if reader.metadata else filename,
                'creation_date': reader.metadata.get('/CreationDate', '') if reader.metadata else ''
            }
            
            return {
                'text': text_content.strip(),
                'metadata': metadata,
                'format': 'pdf'
            }
            
        except Exception as e:
            logger.error(f"PDF processing error: {e}")
            raise HTTPException(status_code=400, detail=f"PDF processing failed: {str(e)}")
    
    def _process_docx(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """Process DOCX document"""
        try:
            text_content = docx2txt.process(io.BytesIO(file_content))
            
            metadata = {
                'format': 'docx',
                'filename': filename,
                'extracted_text_length': len(text_content)
            }
            
            return {
                'text': text_content.strip(),
                'metadata': metadata,
                'format': 'docx'
            }
            
        except Exception as e:
            logger.error(f"DOCX processing error: {e}")
            raise HTTPException(status_code=400, detail=f"DOCX processing failed: {str(e)}")
    
    def _process_doc(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """Process DOC document (legacy format)"""
        try:
            # For DOC files, we'll use a basic text extraction
            # In production, consider using python-docx2txt or antiword
            text_content = "DOC format processing requires additional libraries"
            
            metadata = {
                'format': 'doc',
                'filename': filename,
                'note': 'Legacy DOC format - limited processing'
            }
            
            return {
                'text': text_content,
                'metadata': metadata,
                'format': 'doc'
            }
            
        except Exception as e:
            logger.error(f"DOC processing error: {e}")
            raise HTTPException(status_code=400, detail=f"DOC processing failed: {str(e)}")
    
    def _process_excel(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """Process Excel document"""
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            
            text_content = ""
            sheets_data = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_text = f"\n--- Sheet: {sheet_name} ---\n"
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        sheet_text += row_text + "\n"
                
                sheets_data.append({
                    'sheet_name': sheet_name,
                    'text': sheet_text
                })
                text_content += sheet_text
            
            metadata = {
                'format': 'excel',
                'sheets': [sheet['sheet_name'] for sheet in sheets_data],
                'total_sheets': len(sheets_data),
                'sheets_data': sheets_data
            }
            
            return {
                'text': text_content.strip(),
                'metadata': metadata,
                'format': 'excel'
            }
            
        except Exception as e:
            logger.error(f"Excel processing error: {e}")
            raise HTTPException(status_code=400, detail=f"Excel processing failed: {str(e)}")
    
    def _process_text(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """Process plain text document"""
        try:
            # Detect encoding
            detected = chardet.detect(file_content)
            encoding = detected['encoding'] or 'utf-8'
            
            text_content = file_content.decode(encoding)
            
            metadata = {
                'format': 'text',
                'encoding': encoding,
                'confidence': detected['confidence']
            }
            
            return {
                'text': text_content.strip(),
                'metadata': metadata,
                'format': 'text'
            }
            
        except Exception as e:
            logger.error(f"Text processing error: {e}")
            raise HTTPException(status_code=400, detail=f"Text processing failed: {str(e)}")
    
    def _process_html(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """Process HTML document"""
        try:
            # Detect encoding
            detected = chardet.detect(file_content)
            encoding = detected['encoding'] or 'utf-8'
            
            html_content = file_content.decode(encoding)
            soup = BeautifulSoup(html_content, 'html.parser')
            
            # Extract text content
            text_content = soup.get_text()
            
            # Extract metadata
            title = soup.find('title')
            meta_tags = soup.find_all('meta')
            
            metadata = {
                'format': 'html',
                'title': title.string if title else filename,
                'meta_tags': {tag.get('name', tag.get('property', 'unknown')): tag.get('content', '') 
                             for tag in meta_tags if tag.get('content')},
                'encoding': encoding
            }
            
            return {
                'text': text_content.strip(),
                'metadata': metadata,
                'format': 'html'
            }
            
        except Exception as e:
            logger.error(f"HTML processing error: {e}")
            raise HTTPException(status_code=400, detail=f"HTML processing failed: {str(e)}")
    
    def _process_csv(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """Process CSV document"""
        try:
            # Detect encoding
            detected = chardet.detect(file_content)
            encoding = detected['encoding'] or 'utf-8'
            
            csv_content = file_content.decode(encoding)
            
            # Convert CSV to readable text format
            lines = csv_content.strip().split('\n')
            if lines:
                headers = lines[0].split(',')
                text_content = f"CSV Data with columns: {', '.join(headers)}\n\n"
                text_content += csv_content
            else:
                text_content = csv_content
            
            metadata = {
                'format': 'csv',
                'total_rows': len(lines),
                'headers': headers if lines else [],
                'encoding': encoding
            }
            
            return {
                'text': text_content.strip(),
                'metadata': metadata,
                'format': 'csv'
            }
            
        except Exception as e:
            logger.error(f"CSV processing error: {e}")
            raise HTTPException(status_code=400, detail=f"CSV processing failed: {str(e)}")
    
    def _count_tokens(self, text: str) -> int:
        """Count tokens in text"""
        if self.tokenizer:
            try:
                return len(self.tokenizer.encode(text))
            except Exception:
                pass
        
        # Fallback: approximate token count
        return len(text.split()) * 1.3  # rough approximation
    
    def _semantic_chunking(self, text: str, config: ChunkingConfig) -> List[Dict[str, Any]]:
        """Semantic chunking based on sentences and paragraphs"""
        chunks = []
        paragraphs = text.split('\n\n')
        current_chunk = ""
        current_start = 0
        chunk_index = 0
        
        for para in paragraphs:
            para = para.strip()
            if not para:
                continue
            
            # If adding this paragraph would exceed chunk size
            if len(current_chunk) + len(para) > config.chunk_size and current_chunk:
                # Create chunk from current content
                if len(current_chunk) >= config.min_chunk_size:
                    chunks.append({
                        'chunk_id': f"chunk_{chunk_index}",
                        'text': current_chunk.strip(),
                        'chunk_index': chunk_index,
                        'start_char': current_start,
                        'end_char': current_start + len(current_chunk),
                        'word_count': len(current_chunk.split()),
                        'token_count': int(self._count_tokens(current_chunk))
                    })
                    chunk_index += 1
                
                # Start new chunk with overlap
                if config.chunk_overlap > 0 and current_chunk:
                    overlap_text = current_chunk[-config.chunk_overlap:]
                    current_start = current_start + len(current_chunk) - config.chunk_overlap
                    current_chunk = overlap_text + "\n\n" + para
                else:
                    current_start = current_start + len(current_chunk)
                    current_chunk = para
            else:
                # Add paragraph to current chunk
                if current_chunk:
                    current_chunk += "\n\n" + para
                else:
                    current_chunk = para
        
        # Add final chunk
        if current_chunk and len(current_chunk) >= config.min_chunk_size:
            chunks.append({
                'chunk_id': f"chunk_{chunk_index}",
                'text': current_chunk.strip(),
                'chunk_index': chunk_index,
                'start_char': current_start,
                'end_char': current_start + len(current_chunk),
                'word_count': len(current_chunk.split()),
                'token_count': int(self._count_tokens(current_chunk))
            })
        
        return chunks
    
    def _fixed_chunking(self, text: str, config: ChunkingConfig) -> List[Dict[str, Any]]:
        """Fixed-size chunking with overlap"""
        chunks = []
        chunk_index = 0
        
        for i in range(0, len(text), config.chunk_size - config.chunk_overlap):
            chunk_text = text[i:i + config.chunk_size]
            
            if len(chunk_text) >= config.min_chunk_size:
                chunks.append({
                    'chunk_id': f"chunk_{chunk_index}",
                    'text': chunk_text,
                    'chunk_index': chunk_index,
                    'start_char': i,
                    'end_char': i + len(chunk_text),
                    'word_count': len(chunk_text.split()),
                    'token_count': int(self._count_tokens(chunk_text))
                })
                chunk_index += 1
        
        return chunks
    
    def _sentence_chunking(self, text: str, config: ChunkingConfig) -> List[Dict[str, Any]]:
        """Sentence-based chunking"""
        sentences = sent_tokenize(text)
        chunks = []
        current_chunk = ""
        current_start = 0
        chunk_index = 0
        
        for sentence in sentences:
            if len(current_chunk) + len(sentence) > config.chunk_size and current_chunk:
                # Create chunk
                if len(current_chunk) >= config.min_chunk_size:
                    chunks.append({
                        'chunk_id': f"chunk_{chunk_index}",
                        'text': current_chunk.strip(),
                        'chunk_index': chunk_index,
                        'start_char': current_start,
                        'end_char': current_start + len(current_chunk),
                        'word_count': len(current_chunk.split()),
                        'token_count': int(self._count_tokens(current_chunk))
                    })
                    chunk_index += 1
                
                # Start new chunk
                current_start = current_start + len(current_chunk)
                current_chunk = sentence
            else:
                current_chunk += " " + sentence if current_chunk else sentence
        
        # Add final chunk
        if current_chunk and len(current_chunk) >= config.min_chunk_size:
            chunks.append({
                'chunk_id': f"chunk_{chunk_index}",
                'text': current_chunk.strip(),
                'chunk_index': chunk_index,
                'start_char': current_start,
                'end_char': current_start + len(current_chunk),
                'word_count': len(current_chunk.split()),
                'token_count': int(self._count_tokens(current_chunk))
            })
        
        return chunks
    
    async def process_document(
        self, 
        file_content: bytes, 
        request: DocumentProcessingRequest
    ) -> DocumentProcessingResponse:
        """Process document and create chunks with metadata"""
        start_time = datetime.now()
        
        try:
            # Detect content type
            content_type = self._detect_content_type(file_content, request.filename)
            
            # Check if format is supported
            if content_type not in self.supported_formats:
                raise HTTPException(
                    status_code=400, 
                    detail=f"Unsupported format: {content_type}"
                )
            
            # Process document based on type
            processed_doc = self.supported_formats[content_type](file_content, request.filename)
            
            # Extract text content
            text_content = processed_doc['text']
            doc_metadata = processed_doc['metadata']
            
            # Generate document ID
            doc_id = str(uuid.uuid4())
            
            # Perform chunking
            chunking_method = {
                'semantic': self._semantic_chunking,
                'fixed': self._fixed_chunking,
                'sentence': self._sentence_chunking
            }.get(request.chunking_config.chunking_strategy, self._semantic_chunking)
            
            raw_chunks = chunking_method(text_content, request.chunking_config)
            
            # Create processed chunks with metadata
            processed_chunks = []
            for chunk_data in raw_chunks:
                processed_chunks.append(ProcessedChunk(
                    chunk_id=f"{doc_id}_{chunk_data['chunk_id']}",
                    text=chunk_data['text'],
                    chunk_index=chunk_data['chunk_index'],
                    start_char=chunk_data['start_char'],
                    end_char=chunk_data['end_char'],
                    word_count=chunk_data['word_count'],
                    token_count=chunk_data['token_count']
                ))
            
            # Prepare permission metadata
            permission_metadata = {
                'access_control_list': request.permission_template.access_control_list,
                'roles': request.permission_template.roles,
                'classification': request.permission_template.classification,
                'department': request.permission_template.department,
                'project_id': request.permission_template.project_id,
                'attributes': request.permission_template.attributes
            }
            
            # Enhanced document metadata
            enhanced_metadata = {
                **doc_metadata,
                'original_filename': request.filename,
                'content_type': content_type,
                'file_size': len(file_content),
                'text_length': len(text_content),
                'processed_by': request.user_id,
                'processing_timestamp': datetime.now().isoformat(),
                'chunking_config': request.chunking_config.dict()
            }
            
            # Processing statistics
            processing_time = (datetime.now() - start_time).total_seconds()
            processing_stats = {
                'processing_time_seconds': processing_time,
                'total_chunks': len(processed_chunks),
                'total_words': sum(chunk.word_count for chunk in processed_chunks),
                'total_tokens': sum(chunk.token_count for chunk in processed_chunks),
                'average_chunk_size': sum(len(chunk.text) for chunk in processed_chunks) / len(processed_chunks) if processed_chunks else 0,
                'chunking_strategy': request.chunking_config.chunking_strategy
            }
            
            return DocumentProcessingResponse(
                doc_id=doc_id,
                filename=request.filename,
                processing_status="completed",
                chunks=processed_chunks,
                metadata=enhanced_metadata,
                permission_metadata=permission_metadata,
                processing_stats=processing_stats,
                created_at=datetime.now().isoformat()
            )
            
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"Document processing error: {e}")
            raise HTTPException(status_code=500, detail=f"Processing failed: {str(e)}")

# Global document processor
document_processor = DocumentProcessor()

@asynccontextmanager
async def lifespan(app: FastAPI):
    """Application lifespan manager"""
    logger.info("🚀 Starting Document Processing Service...")
    logger.info("✅ Document Processing Service ready")
    yield
    logger.info("🔌 Shutting down Document Processing Service...")

# FastAPI app
app = FastAPI(
    title="Document Processing Service",
    description="Advanced document processing with chunking and permission management",
    version="1.0.0",
    lifespan=lifespan
)

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Health check
@app.get("/health")
async def health_check():
    """Health check endpoint"""
    return {
        "status": "healthy",
        "service": "document-processing-service",
        "supported_formats": list(document_processor.supported_formats.keys()),
        "timestamp": datetime.now().isoformat()
    }

# Process document from file upload
@app.post("/api/v1/process/upload", response_model=DocumentProcessingResponse)
async def process_uploaded_document(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    permission_template: str = Form(...),
    chunking_config: Optional[str] = Form(None),
    user_id: str = Form(...)
):
    """Process uploaded document with chunking and permission assignment"""
    logger.info(f"📄 Processing uploaded document: {file.filename} by user: {user_id}")
    
    try:
        # Read file content
        file_content = await file.read()
        
        # Parse permission template
        permission_template_obj = PermissionTemplate.parse_raw(permission_template)
        
        # Parse chunking config
        chunking_config_obj = ChunkingConfig()
        if chunking_config:
            chunking_config_obj = ChunkingConfig.parse_raw(chunking_config)
        
        # Create processing request
        processing_request = DocumentProcessingRequest(
            filename=file.filename,
            content_type=file.content_type,
            permission_template=permission_template_obj,
            chunking_config=chunking_config_obj,
            user_id=user_id
        )
        
        # Process document
        result = await document_processor.process_document(file_content, processing_request)
        
        logger.info(f"✅ Document processed: {result.doc_id}, {len(result.chunks)} chunks created")
        
        return result
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Upload processing error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Process document from binary content
@app.post("/api/v1/process/binary", response_model=DocumentProcessingResponse)
async def process_binary_document(processing_data: Dict[str, Any]):
    """Process document from binary content"""
    try:
        # Extract file content (base64 encoded)
        import base64
        file_content = base64.b64decode(processing_data["file_content"])
        
        # Create processing request
        processing_request = DocumentProcessingRequest(**processing_data["request"])
        
        # Process document
        result = await document_processor.process_document(file_content, processing_request)
        
        logger.info(f"✅ Binary document processed: {result.doc_id}")
        
        return result
        
    except Exception as e:
        logger.error(f"❌ Binary processing error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Get supported formats
@app.get("/api/v1/formats")
async def get_supported_formats():
    """Get list of supported document formats"""
    return {
        "supported_formats": list(document_processor.supported_formats.keys()),
        "format_descriptions": {
            "application/pdf": "PDF documents",
            "application/msword": "Microsoft Word documents (legacy)",
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document": "Microsoft Word documents",
            "application/vnd.ms-excel": "Microsoft Excel documents (legacy)",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "Microsoft Excel documents",
            "text/plain": "Plain text documents",
            "text/html": "HTML documents",
            "text/csv": "CSV files"
        }
    }

# Get default chunking configurations
@app.get("/api/v1/chunking/templates")
async def get_chunking_templates():
    """Get chunking configuration templates"""
    return {
        "templates": {
            "default": ChunkingConfig().dict(),
            "large_documents": ChunkingConfig(
                chunk_size=1500,
                chunk_overlap=300,
                chunking_strategy="semantic"
            ).dict(),
            "technical_documents": ChunkingConfig(
                chunk_size=800,
                chunk_overlap=100,
                chunking_strategy="sentence",
                preserve_paragraphs=True
            ).dict(),
            "legal_documents": ChunkingConfig(
                chunk_size=1200,
                chunk_overlap=250,
                chunking_strategy="semantic",
                preserve_paragraphs=True,
                min_chunk_size=200
            ).dict()
        }
    }

if __name__ == "__main__":
    import uvicorn
    import sys
    port = 5000
    if "--port" in sys.argv:
        port_idx = sys.argv.index("--port")
        if port_idx + 1 < len(sys.argv):
            port = int(sys.argv[port_idx + 1])
    uvicorn.run(app, host="0.0.0.0", port=port)